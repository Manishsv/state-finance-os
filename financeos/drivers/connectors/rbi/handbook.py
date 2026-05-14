"""RBI Handbook of Statistics on Indian States — GSDP + population driver.

Source: RBI publication "Handbook of Statistics on Indian States",
specifically Table 21 (Gross State Domestic Product at Current Prices).

GSDP values are read from the cached XLSX. Population values are NOT
in the Handbook — for v0 we use Census 2011 figures projected forward
with state-specific annual growth rates. Documented as a v0 limitation;
v1 should add a proper population driver (Census + UN projections).

Writes to `budget_metadata` table (not `budget_signals`) — these are
denominators used by other metrics, not budget line items themselves.

Mapping decisions:
- GSDP column values are in ₹ Lakh in the source XLSX → converted to
  Crore (÷100) for storage as `gsdp_inr_crore`.
- Population stored as `population_count` in absolute integer units.
- Year format YYYY-YY is used directly (Handbook column headers already
  use this format).
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence

import openpyxl

from financeos.drivers.connectors.base import (
    BudgetDataSourceDriver,
    DriverFetchError,
)
from financeos.drivers.registries.loader import Registries, load_registries
from financeos.drivers.store.ingestor import Ingestor
from financeos.os.conformance import ConformanceResult

DEFAULT_XLSX_PATH = Path("data/raw/rbi_handbook/gsdp_current_prices_2025.xlsx")
EDITION = "2024-25"  # Handbook edition published Dec 2025 covers data through 2024-25

# --- Population baseline: Census 2011 (post-AP-bifurcation for AP/TG) ---
# Source: Census of India 2011. AP and TG split in June 2014 — values shown
# are reconstructed from district-level Census data.
CENSUS_2011_POPULATION: Dict[str, int] = {
    "AP": 49_386_799,    # Andhra Pradesh (residual after TG carve-out)
    "KA": 61_095_297,    # Karnataka
    "KL": 33_406_061,    # Kerala
    "TN": 72_147_030,    # Tamil Nadu
    "TG": 35_286_757,    # Telangana (carved out from united AP)
}

# Annual compound growth assumed from 2011 onwards. Derived from 2001-2011
# decennial growth rates with a slowdown adjustment (post-2011 fertility has
# dropped further). These are point estimates with acknowledged uncertainty
# of ±1 percentage point per state per decade.
ANNUAL_GROWTH_RATES: Dict[str, float] = {
    "AP": 0.005,    # 0.5%/yr — Andhra has near-replacement fertility
    "KA": 0.010,    # 1.0%/yr — Karnataka still growing, urban migration
    "KL": 0.003,    # 0.3%/yr — Kerala below replacement, ageing
    "TN": 0.005,    # 0.5%/yr — Tamil Nadu fertility well below replacement
    "TG": 0.007,    # 0.7%/yr — Telangana mid-pace, Hyderabad pull
}


def project_population(state: str, fiscal_year: str) -> Optional[int]:
    """Project population from Census 2011 baseline using state growth rate.

    fiscal_year is YYYY-YY; we use the YYYY part as the calendar-year proxy.
    """
    if state not in CENSUS_2011_POPULATION:
        return None
    base = CENSUS_2011_POPULATION[state]
    rate = ANNUAL_GROWTH_RATES[state]
    yyyy = int(fiscal_year[:4])
    years_forward = yyyy - 2011
    return int(base * ((1 + rate) ** years_forward))


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


@dataclass
class _GsdpRow:
    state_name: str
    fiscal_year: str
    gsdp_lakh: float


def parse_gsdp_xlsx(xlsx_path: Path) -> List[_GsdpRow]:
    """Read both T_21 sheets and return a flat list of (state, year, GSDP) rows."""
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows: List[_GsdpRow] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Header layout: row 5 has year columns starting at column 3 (C)
        year_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
        # Years sit in columns 2..N (0-indexed in the tuple, since column A is row 0)
        years: List[Optional[str]] = []
        for v in year_row:
            if isinstance(v, str) and len(v) == 7 and v[4] == "-":
                years.append(v)
            else:
                years.append(None)
        # Data rows start at row 6
        for r in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            state_name = r[1]
            if not state_name or not isinstance(state_name, str):
                continue
            for col_idx, year in enumerate(years):
                if year is None:
                    continue
                v = r[col_idx]
                if v is None or v == "" or v == "-":
                    continue
                try:
                    rows.append(_GsdpRow(state_name=state_name.strip(),
                                         fiscal_year=year, gsdp_lakh=float(v)))
                except (TypeError, ValueError):
                    continue
    wb.close()
    return rows


# Map RBI Handbook's state spellings to ISO codes (some differ slightly from the
# State Finances XLSX — e.g. asterisks marking footnoted states)
HANDBOOK_STATE_NAME_TO_CODE: Dict[str, str] = {
    "Andhra Pradesh": "AP",
    "Arunachal Pradesh": "AR",
    "Assam": "AS",
    "Bihar": "BR",
    "Chhattisgarh": "CT",
    "Goa": "GA",
    "Gujarat": "GJ",
    "Haryana": "HR",
    "Himachal Pradesh": "HP",
    "Jammu & Kashmir": "JK",
    "Jammu & Kashmir*": "JK",
    "Jharkhand": "JH",
    "Karnataka": "KA",
    "Kerala": "KL",
    "Madhya Pradesh": "MP",
    "Maharashtra": "MH",
    "Manipur": "MN",
    "Meghalaya": "ML",
    "Mizoram": "MZ",
    "Nagaland": "NL",
    "Odisha": "OR",
    "Punjab": "PB",
    "Rajasthan": "RJ",
    "Sikkim": "SK",
    "Tamil Nadu": "TN",
    "Telangana": "TG",
    "Tripura": "TR",
    "Uttar Pradesh": "UP",
    "Uttarakhand": "UT",
    "West Bengal": "WB",
    "Delhi": "DL",
    "Puducherry": "PY",
}


class RbiHandbookDriver(BudgetDataSourceDriver):
    """Ingest GSDP (from RBI Handbook XLSX) and population (from Census +
    state-specific growth rates) into the `budget_metadata` table."""

    domain = "rbi_handbook"
    cadence_hours = 8760.0
    produces_assessments = False
    signal_names = ["gsdp_inr_crore", "population_count"]
    estimate_types = ["ACT"]   # GSDP is published as actual (NSO releases)
    data_sources = [
        "RBI Handbook of Statistics on Indian States, Table 21 (GSDP at Current Prices)",
        "Census of India 2011 + state-specific growth-rate projections (population)",
    ]
    source_id_template = "rbi.handbook." + EDITION

    def __init__(self, ingestor: Ingestor,
                 xlsx_path: Optional[Path] = None,
                 registries: Optional[Registries] = None):
        self.ingestor = ingestor
        self.xlsx_path = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX_PATH
        self.registries = registries or load_registries()

    def conformance_check(self) -> ConformanceResult:
        failures: List[str] = []
        if not self.xlsx_path.exists():
            failures.append(f"Handbook GSDP XLSX not cached at {self.xlsx_path}.")
        return ConformanceResult(ok=not failures, failures=failures, warnings=[])

    def fetch(self,
              states: Sequence[str],
              fiscal_years: Sequence[str],
              force: bool = False) -> int:
        if not self.xlsx_path.exists():
            raise DriverFetchError(f"Handbook XLSX missing: {self.xlsx_path}")

        wanted_states = set(states)
        wanted_years = set(fiscal_years)

        # GSDP rows from XLSX
        rows_written = 0
        ingested_at = _now_iso()
        source_id = self.source_id_template

        gsdp_rows = parse_gsdp_xlsx(self.xlsx_path)
        for r in gsdp_rows:
            code = HANDBOOK_STATE_NAME_TO_CODE.get(r.state_name)
            if code is None or code not in wanted_states:
                continue
            if r.fiscal_year not in wanted_years:
                continue
            gsdp_crore = r.gsdp_lakh / 100.0  # Lakh → Crore
            self.ingestor.conn.execute(
                """INSERT OR REPLACE INTO budget_metadata
                   (state, fiscal_year, metric, value, unit, source_id, ingested_at)
                   VALUES (?, ?, 'gsdp_inr_crore', ?, 'INR_CRORE', ?, ?)""",
                (code, r.fiscal_year, gsdp_crore, source_id, ingested_at),
            )
            rows_written += 1

        # Population: project Census 2011 forward for every (state, year) pair
        for state in wanted_states:
            for fy in wanted_years:
                pop = project_population(state, fy)
                if pop is None:
                    continue
                self.ingestor.conn.execute(
                    """INSERT OR REPLACE INTO budget_metadata
                       (state, fiscal_year, metric, value, unit, source_id, ingested_at)
                       VALUES (?, ?, 'population_count', ?, 'COUNT',
                               'census2011.projected', ?)""",
                    (state, fy, float(pop), ingested_at),
                )
                rows_written += 1

        return rows_written
